# -*- coding: utf-8 -*-
"""
더술닷컴(thesool.com) 전통주 정보 크롤러.

pip install requests beautifulsoup4 openpyxl

사용 예:
    python sool_crawler.py                 # 기본: 20개 시범 수집
    python sool_crawler.py --limit 0       # 전체 수집 (~1,251개)
    python sool_crawler.py --limit 50 --output my.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from typing import Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BASE = "https://thesool.com"
LIST_URL = BASE + "/front/find/M000000082/list.do?pageIndex={page}"
VIEW_URL = BASE + "/front/find/M000000082/view.do?productId={pid}"
IMAGE_URL = BASE + "/common/imageView.do?targetId={pid}&targetNm=PRODUCT"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "ko,en;q=0.8",
}

COLUMNS = [
    "productId",
    "제품명",
    "종류",
    "원재료",
    "알콜도수",
    "용량",
    "수상내역",
    "기타",
    "태그",
    "제품소개",
    "어울리는음식",
    "양조장명",
    "주소",
    "홈페이지",
    "문의",
    "이미지URL",
    "상세페이지URL",
    "parse_error",
]

PRODUCT_ID_RE = re.compile(r"move\('(PR\d{8})'\)")


def fetch(url: str, retries: int = 3, timeout: int = 15) -> str:
    last_exc: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=timeout)
            resp.raise_for_status()
            resp.encoding = resp.apparent_encoding or "utf-8"
            return resp.text
        except Exception as exc:
            last_exc = exc
            if attempt < retries:
                time.sleep(1.0 * attempt)
    raise RuntimeError(f"GET {url} 실패: {last_exc}")


def get_product_ids(page_index: int) -> list[str]:
    html = fetch(LIST_URL.format(page=page_index))
    ids: list[str] = []
    seen: set[str] = set()
    for match in PRODUCT_ID_RE.finditer(html):
        pid = match.group(1)
        if pid not in seen:
            seen.add(pid)
            ids.append(pid)
    return ids


def _text(node) -> str:
    if node is None:
        return ""
    return node.get_text(" ", strip=True).replace("\xa0", " ").strip()


def _info_list_map(soup: BeautifulSoup) -> dict[str, str]:
    result: dict[str, str] = {}
    ul = soup.select_one("div.product-view ul.info-list")
    if ul is None:
        return result
    for li in ul.select("li"):
        label_node = li.find("strong")
        value_node = li.find("span")
        if not label_node:
            continue
        label = _text(label_node)
        value = _text(value_node)
        result[label] = value
    return result


def _place_map(soup: BeautifulSoup) -> dict[str, str]:
    result: dict[str, str] = {}
    place = soup.select_one("div.product-view dd.place")
    if place is None:
        return result
    for li in place.select("li"):
        label_node = li.find("strong")
        if not label_node:
            continue
        label = _text(label_node)
        # 홈페이지는 <a> 링크의 href 우선
        a = li.find("a")
        if label == "홈페이지" and a and a.get("href"):
            result[label] = a.get("href").strip()
            continue
        # 전체 li 텍스트에서 라벨 제거
        li_text = _text(li)
        if li_text.startswith(label):
            li_text = li_text[len(label):].strip()
        result[label] = li_text
    return result


def parse_detail(product_id: str) -> dict[str, str]:
    row: dict[str, str] = {col: "" for col in COLUMNS}
    row["productId"] = product_id
    row["상세페이지URL"] = VIEW_URL.format(pid=product_id)
    row["이미지URL"] = IMAGE_URL.format(pid=product_id)

    try:
        html = fetch(VIEW_URL.format(pid=product_id))
        soup = BeautifulSoup(html, "html.parser")

        row["제품명"] = _text(soup.select_one("div.product-view dt.subject"))

        info = _info_list_map(soup)
        row["종류"] = info.get("종류", "")
        row["원재료"] = info.get("원재료", "")
        row["알콜도수"] = info.get("알콜도수", "")
        row["용량"] = info.get("용량", "")
        row["수상내역"] = info.get("수상내역", "")
        row["기타"] = info.get("기타", "")

        tags = [
            _text(s)
            for s in soup.select("div.product-view div.tag span")
            if _text(s)
        ]
        row["태그"] = ", ".join(tags)

        row["제품소개"] = _text(soup.select_one("div.product-view dd.intro div.text"))
        row["어울리는음식"] = _text(soup.select_one("div.product-view dd.food div.text"))

        place = _place_map(soup)
        row["양조장명"] = place.get("양조장명", "")
        row["주소"] = place.get("주소", "")
        row["홈페이지"] = place.get("홈페이지", "")
        row["문의"] = place.get("문의", "")
    except Exception as exc:
        row["parse_error"] = f"{type(exc).__name__}: {exc}"

    return row


def write_excel(rows: list[dict[str, str]], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "products"

    ws.append(COLUMNS)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    url_cols = {"이미지URL", "상세페이지URL", "홈페이지"}
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])
        excel_row_idx = ws.max_row
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            if col_name in url_cols:
                value = row.get(col_name, "")
                if value and value.startswith(("http://", "https://")):
                    cell = ws.cell(row=excel_row_idx, column=col_idx)
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")

    widths = {
        "productId": 14,
        "제품명": 28,
        "종류": 14,
        "원재료": 40,
        "알콜도수": 10,
        "용량": 12,
        "수상내역": 30,
        "기타": 20,
        "태그": 24,
        "제품소개": 60,
        "어울리는음식": 40,
        "양조장명": 24,
        "주소": 40,
        "홈페이지": 32,
        "문의": 18,
        "이미지URL": 40,
        "상세페이지URL": 40,
        "parse_error": 30,
    }
    for idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col_name, 18)

    ws.freeze_panes = "A2"
    wb.save(path)


def crawl(limit: int, start_page: int, end_page: int, delay: float) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    collected = 0
    for page in range(start_page, end_page + 1):
        ids = get_product_ids(page)
        if not ids:
            print(f"[경고] page {page} 에서 productId를 찾지 못했습니다. HTML 구조가 변경되었을 수 있습니다.", file=sys.stderr)
            if page == start_page:
                raise SystemExit(1)
            break
        for pid in ids:
            if limit and collected >= limit:
                return rows
            row = parse_detail(pid)
            rows.append(row)
            collected += 1
            name = row.get("제품명") or "(이름 없음)"
            total_label = str(limit) if limit else "전체"
            err = f" [ERR] {row['parse_error']}" if row.get("parse_error") else ""
            print(f"[{collected}/{total_label}] page={page} {pid} {name}{err}")
            time.sleep(delay)
        time.sleep(delay)
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="더술닷컴 전통주 크롤러")
    parser.add_argument("--limit", type=int, default=20, help="수집 개수 (0=전체). 기본 20.")
    parser.add_argument("--start-page", type=int, default=1, help="시작 페이지 (기본 1)")
    parser.add_argument("--end-page", type=int, default=126, help="끝 페이지 (기본 126)")
    parser.add_argument("--delay", type=float, default=0.7, help="요청 간 지연(초), 기본 0.7")
    parser.add_argument("--output", default="sool_products.xlsx", help="출력 엑셀 경로")
    args = parser.parse_args()

    print(f"수집 시작: limit={args.limit or '전체'}, pages={args.start_page}-{args.end_page}, delay={args.delay}s")
    rows = crawl(args.limit, args.start_page, args.end_page, args.delay)
    print(f"총 {len(rows)}건 수집 완료. 엑셀 저장 중 -> {args.output}")
    write_excel(rows, args.output)
    print("완료.")


if __name__ == "__main__":
    main()
